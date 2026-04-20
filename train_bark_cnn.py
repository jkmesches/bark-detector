#!/usr/bin/env python3
"""
Train a small CNN on labeled bark recordings.

Reads bark_labels.xlsx, computes log-mel spectrograms, runs stratified
5-fold CV for an aggregate confusion matrix, then trains a final model
on all data and saves it for the labeling UI to consume.

Artifacts:
  bark_cnn.pt            — final model state_dict
  bark_cnn_meta.json     — audio/mel params + class list
  confusion_matrix.png   — OOF confusion matrix over folds
  training_report.txt    — per-fold metrics + aggregate numbers
"""

from __future__ import annotations

import json
from dataclasses import dataclass, asdict
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import torch
import torch.nn as nn
import torch.nn.functional as F
from openpyxl import load_workbook
from sklearn.metrics import confusion_matrix, classification_report
from sklearn.model_selection import StratifiedKFold

import librosa

PROJECT = Path(__file__).resolve().parent
REC_DIR = PROJECT / "recordings"
XLSX = PROJECT / "bark_labels.xlsx"
MODEL_PATH = PROJECT / "bark_cnn.pt"
META_PATH = PROJECT / "bark_cnn_meta.json"
CM_PATH = PROJECT / "confusion_matrix.png"
REPORT_PATH = PROJECT / "training_report.txt"


@dataclass
class AudioCfg:
    sample_rate: int = 48000
    clip_seconds: float = 3.0
    n_fft: int = 1024
    hop_length: int = 256
    n_mels: int = 64
    f_min: float = 50.0
    f_max: float = 20000.0

    @property
    def n_samples(self) -> int:
        return int(self.sample_rate * self.clip_seconds)


def load_clip(path: Path, cfg: AudioCfg) -> np.ndarray:
    y, _ = librosa.load(str(path), sr=cfg.sample_rate, mono=True)
    n = cfg.n_samples
    if len(y) >= n:
        y = y[:n]
    else:
        y = np.pad(y, (0, n - len(y)), mode="constant")
    return y.astype(np.float32)


def logmel(y: np.ndarray, cfg: AudioCfg) -> np.ndarray:
    S = librosa.feature.melspectrogram(
        y=y,
        sr=cfg.sample_rate,
        n_fft=cfg.n_fft,
        hop_length=cfg.hop_length,
        n_mels=cfg.n_mels,
        fmin=cfg.f_min,
        fmax=cfg.f_max,
        power=2.0,
    )
    return librosa.power_to_db(S, ref=np.max).astype(np.float32)


def read_labels() -> tuple[list[Path], np.ndarray]:
    wb = load_workbook(XLSX)
    ws = wb.active
    paths: list[Path] = []
    labels: list[int] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[2]:
            continue
        filename, label = row[2], row[4]
        if label not in ("YES", "NO"):
            continue
        p = REC_DIR / filename
        if not p.exists():
            continue
        paths.append(p)
        labels.append(1 if label == "YES" else 0)
    return paths, np.array(labels, dtype=np.int64)


def gpu_spec_augment(x: torch.Tensor) -> torch.Tensor:
    """SpecAugment on a (B, 1, mel, T) tensor already on GPU. In-place-style, new tensor."""
    x = x.clone()
    B, _, n_mels, n_frames = x.shape
    fill = x.amin(dim=(2, 3), keepdim=True)
    for _ in range(2):
        w = torch.randint(1, max(2, n_mels // 8), (B,), device=x.device)
        s = (torch.rand(B, device=x.device) * (n_mels - w).clamp(min=1).float()).long()
        active = torch.rand(B, device=x.device) < 0.5
        for b in range(B):
            if active[b]:
                x[b, :, s[b]:s[b] + w[b], :] = fill[b]
    for _ in range(2):
        w = torch.randint(1, max(2, n_frames // 8), (B,), device=x.device)
        s = (torch.rand(B, device=x.device) * (n_frames - w).clamp(min=1).float()).long()
        active = torch.rand(B, device=x.device) < 0.5
        for b in range(B):
            if active[b]:
                x[b, :, :, s[b]:s[b] + w[b]] = fill[b]
    shift = torch.randint(-n_frames // 10, n_frames // 10 + 1, (B,), device=x.device)
    for b in range(B):
        if shift[b] != 0:
            x[b] = torch.roll(x[b], shifts=int(shift[b]), dims=-1)
    return x


class BarkCNN(nn.Module):
    def __init__(self):
        super().__init__()
        self.net = nn.Sequential(
            nn.Conv2d(1, 16, 3, padding=1), nn.BatchNorm2d(16), nn.ReLU(),
            nn.MaxPool2d(2),
            nn.Conv2d(16, 32, 3, padding=1), nn.BatchNorm2d(32), nn.ReLU(),
            nn.MaxPool2d(2),
            nn.Conv2d(32, 64, 3, padding=1), nn.BatchNorm2d(64), nn.ReLU(),
            nn.MaxPool2d(2),
            nn.Conv2d(64, 64, 3, padding=1), nn.BatchNorm2d(64), nn.ReLU(),
            nn.AdaptiveAvgPool2d(1),
            nn.Flatten(),
            nn.Dropout(0.3),
            nn.Linear(64, 1),
        )

    def forward(self, x):
        return self.net(x).squeeze(-1)


def normalize_specs(specs: np.ndarray, mean: float | None = None, std: float | None = None):
    if mean is None:
        mean = float(specs.mean())
        std = float(specs.std() + 1e-8)
    return (specs - mean) / std, mean, std


def train_one(model, X_tr, y_tr, X_va, y_va, device, epochs, pos_weight,
              batch_size=64, augment=True, lr=1e-3, wd=1e-4, patience=12):
    """Fully-on-GPU training: specs live on device, shuffle via index perm."""
    opt = torch.optim.AdamW(model.parameters(), lr=lr, weight_decay=wd)
    sched = torch.optim.lr_scheduler.CosineAnnealingLR(opt, T_max=epochs)
    crit = nn.BCEWithLogitsLoss(pos_weight=torch.tensor([pos_weight], device=device))
    n_tr = X_tr.shape[0]
    best_val = float("inf")
    best_state = None
    bad = 0

    for ep in range(epochs):
        model.train()
        perm = torch.randperm(n_tr, device=device)
        for i in range(0, n_tr, batch_size):
            idx = perm[i:i + batch_size]
            xb = X_tr[idx]
            yb = y_tr[idx]
            if augment:
                xb = gpu_spec_augment(xb)
            opt.zero_grad()
            logits = model(xb)
            loss = crit(logits, yb)
            loss.backward()
            opt.step()
        sched.step()

        if X_va is None:
            continue
        model.eval()
        with torch.no_grad():
            vlogits = model(X_va)
            vloss = float(crit(vlogits, y_va).item())
        if vloss < best_val - 1e-4:
            best_val = vloss
            best_state = {k: v.detach().cpu().clone() for k, v in model.state_dict().items()}
            bad = 0
        else:
            bad += 1
            if bad >= patience:
                break
    if best_state is not None:
        model.load_state_dict(best_state)
    return model


def predict_probs(model, X: torch.Tensor) -> np.ndarray:
    model.eval()
    with torch.no_grad():
        logits = model(X)
        return torch.sigmoid(logits).cpu().numpy()


def plot_confusion(cm: np.ndarray, path: Path, title: str):
    fig, ax = plt.subplots(figsize=(5.2, 4.6))
    im = ax.imshow(cm, cmap="Blues")
    ax.set_xticks([0, 1], ["NO", "YES"])
    ax.set_yticks([0, 1], ["NO", "YES"])
    ax.set_xlabel("Predicted")
    ax.set_ylabel("Actual")
    ax.set_title(title)
    for i in range(2):
        for j in range(2):
            ax.text(j, i, str(cm[i, j]), ha="center", va="center",
                    color="white" if cm[i, j] > cm.max() / 2 else "black")
    fig.colorbar(im, ax=ax)
    fig.tight_layout()
    fig.savefig(path, dpi=120)
    plt.close(fig)


def main():
    cfg = AudioCfg()
    print(f"[config] {cfg}")

    paths, y = read_labels()
    print(f"[data] {len(paths)} labeled clips ({int((y==1).sum())} YES / {int((y==0).sum())} NO)")

    print("[feat] computing log-mel spectrograms...")
    specs = np.stack([logmel(load_clip(p, cfg), cfg) for p in paths])
    print(f"[feat] specs shape: {specs.shape} (N, n_mels, T)")

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"[device] {device}  {torch.cuda.get_device_name(0) if device.type == 'cuda' else ''}")

    n_splits = 5
    skf = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=42)
    oof_probs = np.zeros(len(y), dtype=np.float32)

    lines: list[str] = []
    for fold, (tr, va) in enumerate(skf.split(specs, y)):
        X_tr_np, m, s = normalize_specs(specs[tr])
        X_va_np, _, _ = normalize_specs(specs[va], m, s)
        X_tr = torch.from_numpy(X_tr_np).unsqueeze(1).to(device)
        X_va = torch.from_numpy(X_va_np).unsqueeze(1).to(device)
        y_tr = torch.from_numpy(y[tr].astype(np.float32)).to(device)
        y_va = torch.from_numpy(y[va].astype(np.float32)).to(device)

        n_pos = int((y[tr] == 1).sum())
        n_neg = int((y[tr] == 0).sum())
        pw = n_neg / max(n_pos, 1)

        model = BarkCNN().to(device)
        train_one(model, X_tr, y_tr, X_va, y_va, device, epochs=60, pos_weight=pw)
        probs = predict_probs(model, X_va)
        oof_probs[va] = probs

        preds = (probs >= 0.5).astype(int)
        y_va_np = y[va]
        cm = confusion_matrix(y_va_np, preds, labels=[0, 1])
        acc = float((preds == y_va_np).mean())
        line = f"[fold {fold+1}/{n_splits}] acc={acc:.3f}  n_train={len(tr)} n_val={len(va)}  cm={cm.tolist()}"
        print(line, flush=True)
        lines.append(line)

    oof_preds = (oof_probs >= 0.5).astype(int)
    cm = confusion_matrix(y, oof_preds, labels=[0, 1])
    acc = float((oof_preds == y).mean())
    report = classification_report(y, oof_preds, labels=[0, 1], target_names=["NO", "YES"], digits=3)
    print(f"\n[aggregate OOF] acc={acc:.3f}")
    print(f"[aggregate OOF] confusion matrix:\n{cm}")
    print(report)
    plot_confusion(cm, CM_PATH, title=f"OOF confusion (n={len(y)}, acc={acc:.3f})")
    print(f"[out] wrote {CM_PATH}")

    print("\n[final] training on all data for deployment...", flush=True)
    X_all_np, mean_all, std_all = normalize_specs(specs)
    X_all = torch.from_numpy(X_all_np).unsqueeze(1).to(device)
    y_all = torch.from_numpy(y.astype(np.float32)).to(device)
    n_pos = int((y == 1).sum())
    n_neg = int((y == 0).sum())
    pw = n_neg / max(n_pos, 1)
    final = BarkCNN().to(device)
    train_one(final, X_all, y_all, None, None, device, epochs=60, pos_weight=pw, patience=9999)

    torch.save(final.state_dict(), MODEL_PATH)
    meta = {
        "audio": asdict(cfg),
        "normalize": {"mean": mean_all, "std": std_all},
        "classes": ["NO", "YES"],
        "n_train": int(len(y)),
        "oof_accuracy": acc,
        "oof_confusion_matrix": cm.tolist(),
        "default_threshold": 0.5,
    }
    META_PATH.write_text(json.dumps(meta, indent=2))
    print(f"[out] wrote {MODEL_PATH}")
    print(f"[out] wrote {META_PATH}")

    with REPORT_PATH.open("w") as f:
        f.write("\n".join(lines))
        f.write(f"\n\n[aggregate OOF] acc={acc:.3f}\n")
        f.write(f"confusion matrix (rows=actual NO/YES, cols=pred NO/YES):\n{cm}\n\n")
        f.write(report)
    print(f"[out] wrote {REPORT_PATH}")


if __name__ == "__main__":
    main()
