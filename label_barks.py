#!/usr/bin/env python3
"""
Simple labeling UI for bark recordings.

Keys:
  Right Arrow / Y  -> YES (real bark)
  Left  Arrow / N  -> NO  (false positive)
  Down  Arrow / E  -> EXCLUDE (skip this clip from training/testing)
  Space            -> replay audio
  Backspace        -> undo last label
  Esc              -> quit
"""

import argparse
import json
import random
import re
import subprocess
import sys
import threading
import wave
from datetime import datetime
from pathlib import Path

import numpy as np
import sounddevice as sd
import tkinter as tk
from tkinter import ttk
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from openpyxl import Workbook, load_workbook

PROJECT = Path(__file__).resolve().parent
REC_DIR = PROJECT / "recordings"
UNIDENTIFIED_DIR = REC_DIR / "unidentified"
LOG_FILE = PROJECT / "bark_detector.log"
XLSX = PROJECT / "bark_labels.xlsx"
MODEL_PATH = PROJECT / "bark_cnn.pt"
META_PATH = PROJECT / "bark_cnn_meta.json"
THRESHOLD_PATH = PROJECT / "bark_cnn_threshold.json"
SCORES_PATH = PROJECT / "bark_scores.json"
TRAIN_SCRIPT = PROJECT / "train_bark_cnn.py"

COLUMNS = ["bark_id", "timestamp", "filename", "peak_dbfs", "label"]
MQTT_RE = re.compile(r"Published to homeassistant/mic/dog: (\{.*\})")
EXCLUDE_DATES = {"20260407"}


def parse_log_dbfs() -> dict[str, float]:
    """Map recording filename -> peak_rms_dbfs from the detector log."""
    out: dict[str, float] = {}
    if not LOG_FILE.exists():
        return out
    with LOG_FILE.open("r", errors="ignore") as f:
        for line in f:
            m = MQTT_RE.search(line)
            if not m:
                continue
            try:
                d = json.loads(m.group(1))
                rec = d.get("recording")
                db = d.get("peak_rms_dbfs")
                if rec and db is not None:
                    out[rec] = round(float(db), 2)
            except (json.JSONDecodeError, TypeError, ValueError):
                continue
    return out


def load_wav(path: Path) -> tuple[np.ndarray, int]:
    with wave.open(str(path), "rb") as w:
        n = w.getnframes()
        sr = w.getframerate()
        ch = w.getnchannels()
        sw = w.getsampwidth()
        raw = w.readframes(n)
    dtype = {1: np.int8, 2: np.int16, 4: np.int32}[sw]
    arr = np.frombuffer(raw, dtype=dtype)
    if ch > 1:
        arr = arr.reshape(-1, ch).mean(axis=1)
    return arr.astype(np.float32) / float(np.iinfo(dtype).max), sr


def bark_id_from_name(name: str) -> str:
    stem = Path(name).stem
    return stem.removeprefix("bark_")


def timestamp_from_name(name: str) -> str:
    bid = bark_id_from_name(name)
    try:
        return datetime.strptime(bid, "%Y%m%d_%H%M%S").isoformat()
    except ValueError:
        return ""


class BarkPredictor:
    """Lazy-load a trained CNN and score WAVs. Safe to construct without a model file."""

    def __init__(self):
        self.available = False
        self.error: str | None = None
        self.meta: dict | None = None
        self._model = None
        self._torch = None
        self._librosa = None
        self._device = None
        self._try_load()

    def _try_load(self):
        if not MODEL_PATH.exists() or not META_PATH.exists():
            self.error = "no model on disk — run train_bark_cnn.py"
            return
        try:
            import torch
            import librosa
        except ImportError as e:
            self.error = f"missing deps: {e}"
            return
        try:
            meta = json.loads(META_PATH.read_text())
            from train_bark_cnn import BarkCNN
            device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
            model = BarkCNN().to(device)
            state = torch.load(MODEL_PATH, map_location=device, weights_only=True)
            model.load_state_dict(state)
            model.eval()
            self._torch = torch
            self._librosa = librosa
            self._device = device
            self._model = model
            self.meta = meta
            self.available = True
        except Exception as e:
            self.error = f"load failed: {e}"

    def _logmel(self, wav_path: Path) -> np.ndarray:
        cfg = self.meta["audio"]
        norm = self.meta["normalize"]
        y, _ = self._librosa.load(str(wav_path), sr=cfg["sample_rate"], mono=True)
        n = int(cfg["sample_rate"] * cfg["clip_seconds"])
        y = y[:n] if len(y) >= n else np.pad(y, (0, n - len(y)))
        S = self._librosa.feature.melspectrogram(
            y=y.astype(np.float32),
            sr=cfg["sample_rate"],
            n_fft=cfg["n_fft"],
            hop_length=cfg["hop_length"],
            n_mels=cfg["n_mels"],
            fmin=cfg["f_min"],
            fmax=cfg["f_max"],
            power=2.0,
        )
        S = self._librosa.power_to_db(S, ref=np.max).astype(np.float32)
        return (S - norm["mean"]) / norm["std"]

    def score(self, wav_path: Path) -> float | None:
        if not self.available:
            return None
        S = self._logmel(wav_path)
        x = self._torch.from_numpy(S).unsqueeze(0).unsqueeze(0).to(self._device)
        with self._torch.no_grad():
            logit = self._model(x).item()
        return 1.0 / (1.0 + float(np.exp(-logit)))

    def score_batch(self, paths: list[Path], batch_size: int = 64) -> list[float]:
        """Score a list of files. Feature extraction is CPU-bound; batching the forward pass on GPU."""
        if not self.available:
            return []
        out: list[float] = []
        for i in range(0, len(paths), batch_size):
            batch_paths = paths[i:i + batch_size]
            specs = np.stack([self._logmel(p) for p in batch_paths])
            x = self._torch.from_numpy(specs).unsqueeze(1).to(self._device)
            with self._torch.no_grad():
                logits = self._model(x).detach().cpu().numpy()
            probs = 1.0 / (1.0 + np.exp(-logits))
            out.extend(probs.tolist())
        return out


def load_threshold(default: float) -> float:
    if THRESHOLD_PATH.exists():
        try:
            return float(json.loads(THRESHOLD_PATH.read_text()).get("threshold", default))
        except (json.JSONDecodeError, ValueError):
            pass
    return default


def save_threshold(value: float):
    THRESHOLD_PATH.write_text(json.dumps({"threshold": round(float(value), 3)}))


def _model_version() -> float:
    return MODEL_PATH.stat().st_mtime if MODEL_PATH.exists() else 0.0


def load_scores() -> dict[str, float]:
    if not SCORES_PATH.exists():
        return {}
    try:
        data = json.loads(SCORES_PATH.read_text())
    except json.JSONDecodeError:
        return {}
    if float(data.get("model_version", 0.0)) != _model_version():
        return {}
    scores = data.get("scores", {})
    return {k: float(v) for k, v in scores.items()}


def save_scores(scores: dict[str, float]):
    SCORES_PATH.write_text(json.dumps({
        "model_version": _model_version(),
        "scores": {k: round(float(v), 4) for k, v in scores.items()},
    }))


class LabelerWorkbook:
    def __init__(self, path: Path):
        self.path = path
        if path.exists():
            self.wb = load_workbook(path)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "labels"
            self.ws.append(COLUMNS)
            self.wb.save(path)
        self._labeled = self._read_labeled_set()

    def _read_labeled_set(self) -> set[str]:
        labeled = set()
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if row and row[2]:
                labeled.add(row[2])
        return labeled

    def is_labeled(self, filename: str) -> bool:
        return filename in self._labeled

    def append(self, bark_id, ts, filename, db, label):
        self.ws.append([bark_id, ts, filename, db, label])
        self._labeled.add(filename)
        self.wb.save(self.path)

    def pop_last(self) -> tuple[str, str] | None:
        max_row = self.ws.max_row
        if max_row < 2:
            return None
        filename = self.ws.cell(row=max_row, column=3).value
        label = self.ws.cell(row=max_row, column=5).value
        self.ws.delete_rows(max_row, 1)
        if filename in self._labeled:
            self._labeled.discard(filename)
        self.wb.save(self.path)
        return filename, label


class App:
    def __init__(self, root: tk.Tk, files: list[Path], db_map: dict[str, float],
                 wb: LabelerWorkbook, predictor: BarkPredictor):
        self.root = root
        self.files = files
        self.db_map = db_map
        self.wb = wb
        self.predictor = predictor
        self.unlabeled = [p for p in files if not wb.is_labeled(p.name)]
        self.queue: list[Path] = []
        self.total = len(files)
        self.idx_in_queue = 0
        self.current: Path | None = None
        self.current_audio: tuple[np.ndarray, int] | None = None
        self.current_score: float | None = None
        self.score_cache: dict[str, float] = load_scores() if predictor.available else {}
        self.score_min = 0.0
        self.score_max = 1.0
        self.threshold = load_threshold(
            (predictor.meta or {}).get("default_threshold", 0.5) if predictor.available else 0.5
        )

        root.title("Bark Labeler")
        root.geometry("1080x720")

        self.status = tk.Label(root, text="", font=("TkDefaultFont", 11), anchor="w", justify="left")
        self.status.pack(fill="x", padx=8, pady=(8, 0))

        score_row = tk.Frame(root)
        score_row.pack(fill="x", padx=8, pady=(4, 0))
        self.score_label = tk.Label(
            score_row, text="score: —", font=("TkDefaultFont", 14, "bold"), width=28, anchor="w"
        )
        self.score_label.pack(side="left")
        self.pred_label = tk.Label(score_row, text="", font=("TkDefaultFont", 12), anchor="w")
        self.pred_label.pack(side="left", padx=10)

        thr_row = tk.Frame(root)
        thr_row.pack(fill="x", padx=8, pady=(4, 0))
        tk.Label(thr_row, text="threshold:", font=("TkDefaultFont", 10)).pack(side="left")
        self.thr_var = tk.DoubleVar(value=self.threshold)
        self.thr_scale = ttk.Scale(
            thr_row, from_=0.0, to=1.0, orient="horizontal",
            variable=self.thr_var, command=self._on_threshold_change, length=360,
        )
        self.thr_scale.pack(side="left", padx=6)
        self.thr_value = tk.Label(thr_row, text=f"{self.threshold:.2f}", font=("TkDefaultFont", 10), width=5)
        self.thr_value.pack(side="left")

        self.retrain_btn = tk.Button(thr_row, text="Retrain", command=self.retrain)
        self.retrain_btn.pack(side="right")

        filter_row = tk.Frame(root)
        filter_row.pack(fill="x", padx=8, pady=(4, 0))
        tk.Label(filter_row, text="filter score:", font=("TkDefaultFont", 10)).pack(side="left")
        self.min_var = tk.DoubleVar(value=0.0)
        self.max_var = tk.DoubleVar(value=1.0)
        self.min_spin = tk.Spinbox(
            filter_row, from_=0.0, to=1.0, increment=0.05, format="%.2f",
            width=6, textvariable=self.min_var,
        )
        self.min_spin.pack(side="left", padx=(4, 2))
        tk.Label(filter_row, text="to", font=("TkDefaultFont", 10)).pack(side="left")
        self.max_spin = tk.Spinbox(
            filter_row, from_=0.0, to=1.0, increment=0.05, format="%.2f",
            width=6, textvariable=self.max_var,
        )
        self.max_spin.pack(side="left", padx=(2, 6))
        tk.Button(filter_row, text="Apply", command=self.apply_filter).pack(side="left", padx=4)
        tk.Button(filter_row, text="Reset", command=self.reset_filter).pack(side="left", padx=4)
        self.score_all_btn = tk.Button(filter_row, text="Score Unlabeled", command=self.score_all_unlabeled)
        self.score_all_btn.pack(side="left", padx=4)
        self.filter_stats = tk.Label(filter_row, text="", font=("TkDefaultFont", 10))
        self.filter_stats.pack(side="left", padx=10)

        self.fig = Figure(figsize=(10, 4), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.canvas = FigureCanvasTkAgg(self.fig, master=root)
        self.canvas.get_tk_widget().pack(fill="both", expand=True, padx=8, pady=8)

        hint = tk.Label(
            root,
            text="← / N = NO    → / Y = YES    ↓ / E = EXCLUDE    Space = replay    Backspace = undo    Esc = quit",
            font=("TkDefaultFont", 10),
        )
        hint.pack(side="bottom", pady=6)

        btns = tk.Frame(root)
        btns.pack(side="bottom")
        tk.Button(btns, text="NO (←)", width=14, command=lambda: self.label("NO")).pack(side="left", padx=6)
        tk.Button(btns, text="EXCLUDE (↓)", width=14, command=lambda: self.label("EXCLUDE")).pack(side="left", padx=6)
        tk.Button(btns, text="YES (→)", width=14, command=lambda: self.label("YES")).pack(side="left", padx=6)

        root.bind("<Left>", lambda _e: self.label("NO"))
        root.bind("<Right>", lambda _e: self.label("YES"))
        root.bind("<Down>", lambda _e: self.label("EXCLUDE"))
        root.bind("<Key-n>", lambda _e: self.label("NO"))
        root.bind("<Key-N>", lambda _e: self.label("NO"))
        root.bind("<Key-y>", lambda _e: self.label("YES"))
        root.bind("<Key-Y>", lambda _e: self.label("YES"))
        root.bind("<Key-e>", lambda _e: self.label("EXCLUDE"))
        root.bind("<Key-E>", lambda _e: self.label("EXCLUDE"))
        root.bind("<space>", lambda _e: self.play())
        root.bind("<BackSpace>", lambda _e: self.undo())
        root.bind("<Escape>", lambda _e: root.destroy())

        self._rebuild_queue()
        self.show_current()

    def _rebuild_queue(self):
        remaining = [p for p in self.unlabeled if not self.wb.is_labeled(p.name)]
        self.unlabeled = remaining
        if self.score_min <= 0.0 and self.score_max >= 1.0:
            self.queue = list(remaining)
        else:
            self.queue = [
                p for p in remaining
                if p.name in self.score_cache
                and self.score_min <= self.score_cache[p.name] <= self.score_max
            ]
        random.shuffle(self.queue)
        self.idx_in_queue = 0
        self._update_filter_stats()

    def _update_filter_stats(self):
        n_scored = sum(1 for p in self.unlabeled if p.name in self.score_cache)
        self.filter_stats.config(
            text=f"queue: {len(self.queue)}   unlabeled: {len(self.unlabeled)}   scored: {n_scored}"
        )

    def apply_filter(self):
        try:
            lo = max(0.0, min(1.0, float(self.min_var.get())))
            hi = max(0.0, min(1.0, float(self.max_var.get())))
        except (tk.TclError, ValueError):
            self.status.config(text="filter: invalid range")
            return
        if lo > hi:
            lo, hi = hi, lo
        self.score_min, self.score_max = lo, hi
        self.min_var.set(lo)
        self.max_var.set(hi)
        if not self.predictor.available:
            self.status.config(text="filter: model unavailable, showing all")
            self.score_min, self.score_max = 0.0, 1.0
        missing = [p for p in self.unlabeled if p.name not in self.score_cache]
        if missing and self.predictor.available and (lo > 0.0 or hi < 1.0):
            self.status.config(
                text=f"filter: {len(missing)} unlabeled clips have no score — click 'Score Unlabeled' first"
            )
        self._rebuild_queue()
        self.show_current()

    def reset_filter(self):
        self.min_var.set(0.0)
        self.max_var.set(1.0)
        self.score_min, self.score_max = 0.0, 1.0
        self._rebuild_queue()
        self.show_current()

    def score_all_unlabeled(self):
        if not self.predictor.available:
            self.status.config(text="score: model unavailable")
            return
        missing = [p for p in self.unlabeled if p.name not in self.score_cache]
        if not missing:
            self.status.config(text="score: all unlabeled clips already scored")
            self._update_filter_stats()
            return
        self.score_all_btn.config(state="disabled")

        def worker():
            total = len(missing)
            batch = 64
            done = 0
            for i in range(0, total, batch):
                chunk = missing[i:i + batch]
                try:
                    probs = self.predictor.score_batch(chunk, batch_size=batch)
                except Exception as e:
                    self.root.after(0, lambda e=e: self.status.config(text=f"score error: {e}"))
                    break
                for p, s in zip(chunk, probs):
                    self.score_cache[p.name] = float(s)
                done += len(chunk)
                self.root.after(0, lambda d=done, t=total: self.status.config(
                    text=f"scoring unlabeled clips: {d}/{t}"
                ))
            save_scores(self.score_cache)

            def finish():
                self.score_all_btn.config(state="normal")
                self.status.config(text=f"scored {done}/{total} clips")
                self._update_filter_stats()

            self.root.after(0, finish)

        threading.Thread(target=worker, daemon=True).start()

    def labeled_count(self) -> int:
        return self.total - len(self.unlabeled)

    def _on_threshold_change(self, _value):
        self.threshold = float(self.thr_var.get())
        self.thr_value.config(text=f"{self.threshold:.2f}")
        save_threshold(self.threshold)
        self._update_score_display()

    def _update_score_display(self):
        if self.current_score is None:
            if self.predictor.available:
                self.score_label.config(text="score: —", fg="black")
            else:
                self.score_label.config(text=f"score: unavailable ({self.predictor.error})", fg="gray")
            self.pred_label.config(text="")
            return
        s = self.current_score
        pred = "YES" if s >= self.threshold else "NO"
        color = "#0a7f1f" if pred == "YES" else "#b11d1d"
        self.score_label.config(text=f"score: {s:.3f}", fg=color)
        self.pred_label.config(text=f"predicted: {pred}  (threshold {self.threshold:.2f})", fg=color)

    def show_current(self):
        sd.stop()
        if self.idx_in_queue >= len(self.queue):
            self.ax.clear()
            msg = "All done." if self.labeled_count() >= self.total else "Queue empty (filter)."
            self.ax.text(0.5, 0.5, msg, ha="center", va="center", fontsize=20)
            self.ax.set_axis_off()
            self.canvas.draw()
            self.status.config(text=f"Labeled {self.labeled_count()} / {self.total}. Saved to {self.wb.path.name}.")
            self.current = None
            self.current_score = None
            self._update_score_display()
            return

        path = self.queue[self.idx_in_queue]
        self.current = path
        try:
            audio, sr = load_wav(path)
            self.current_audio = (audio, sr)
        except Exception as e:
            self.current_audio = None
            self.ax.clear()
            self.ax.text(0.5, 0.5, f"Cannot load:\n{e}", ha="center", va="center")
            self.ax.set_axis_off()
            self.canvas.draw()
            self.status.config(
                text=f"[q {self.idx_in_queue+1}/{len(self.queue)}] labeled {self.labeled_count()}/{self.total} — {path.name}  (read error)"
            )
            return

        t = np.arange(len(audio)) / sr
        self.ax.clear()
        self.ax.plot(t, audio, linewidth=0.6)
        self.ax.set_xlim(0, t[-1] if len(t) else 1)
        self.ax.set_ylim(-1.05, 1.05)
        self.ax.set_xlabel("Time (s)")
        self.ax.set_ylabel("Amplitude")
        self.ax.grid(alpha=0.3)
        self.fig.tight_layout()
        self.canvas.draw()

        db = self.db_map.get(path.name)
        db_str = f"{db:+.1f} dBFS" if db is not None else "dBFS: n/a"
        self.status.config(
            text=f"[q {self.idx_in_queue+1}/{len(self.queue)}]  labeled {self.labeled_count()}/{self.total}   "
                 f"{path.name}   peak={db_str}   dur={len(audio)/sr:.2f}s"
        )

        cached = self.score_cache.get(path.name)
        self.current_score = cached
        self._update_score_display()
        if self.predictor.available and cached is None:
            try:
                self.current_score = self.predictor.score(path)
                if self.current_score is not None:
                    self.score_cache[path.name] = self.current_score
            except Exception as e:
                print(f"[predict] {path.name}: {e}", file=sys.stderr)
                self.current_score = None
            self._update_score_display()

        self.play()

    def play(self):
        if self.current_audio is None:
            return
        audio, sr = self.current_audio
        sd.stop()
        try:
            sd.play(audio, sr)
        except Exception as e:
            print(f"[play] error: {e}", file=sys.stderr)

    def label(self, value: str):
        if self.current is None:
            return
        path = self.current
        name = path.name
        if value == "EXCLUDE":
            UNIDENTIFIED_DIR.mkdir(exist_ok=True)
            dest = UNIDENTIFIED_DIR / name
            try:
                path.rename(dest)
            except OSError as e:
                print(f"[exclude] could not move {name}: {e}", file=sys.stderr)
                self.status.config(text=f"exclude failed: {e}")
                return
        self.wb.append(
            bark_id_from_name(name),
            timestamp_from_name(name),
            name,
            self.db_map.get(name),
            value,
        )
        self.unlabeled = [p for p in self.unlabeled if p.name != name]
        self.idx_in_queue += 1
        self._update_filter_stats()
        self.show_current()

    def undo(self):
        if self.idx_in_queue > 0:
            self.idx_in_queue -= 1
            last = self.wb.pop_last()
            if last:
                _, prev_label = last
                prev_file = self.queue[self.idx_in_queue]
                if prev_file not in self.unlabeled:
                    self.unlabeled.append(prev_file)
                if prev_label == "EXCLUDE":
                    src = UNIDENTIFIED_DIR / prev_file.name
                    if src.exists():
                        try:
                            src.rename(prev_file)
                        except OSError as e:
                            print(f"[undo] could not restore {prev_file.name}: {e}", file=sys.stderr)
            self._update_filter_stats()
            self.show_current()
        else:
            last = self.wb.pop_last()
            if last:
                filename, prev_label = last
                if prev_label == "EXCLUDE":
                    src = UNIDENTIFIED_DIR / filename
                    dst = REC_DIR / filename
                    if src.exists():
                        try:
                            src.rename(dst)
                        except OSError as e:
                            print(f"[undo] could not restore {filename}: {e}", file=sys.stderr)
                for p in self.files:
                    if p.name == filename:
                        if p not in self.unlabeled:
                            self.unlabeled.append(p)
                        self.queue.insert(0, p)
                        break
                self._update_filter_stats()
                self.show_current()

    def retrain(self):
        if not TRAIN_SCRIPT.exists():
            self.status.config(text="train_bark_cnn.py not found")
            return
        self.retrain_btn.config(state="disabled", text="Retraining…")

        def worker():
            try:
                proc = subprocess.run(
                    [sys.executable, str(TRAIN_SCRIPT)],
                    capture_output=True, text=True, cwd=str(PROJECT),
                )
                ok = proc.returncode == 0
                msg = "retrain complete" if ok else f"retrain failed (exit {proc.returncode})"
                print(proc.stdout)
                if proc.stderr:
                    print(proc.stderr, file=sys.stderr)
            except Exception as e:
                ok = False
                msg = f"retrain error: {e}"

            def finish():
                self.retrain_btn.config(state="normal", text="Retrain")
                self.status.config(text=msg)
                if ok:
                    self.predictor = BarkPredictor()
                    self.score_cache = load_scores() if self.predictor.available else {}
                    self._update_filter_stats()
                    if self.predictor.available:
                        self.threshold = load_threshold(
                            self.predictor.meta.get("default_threshold", 0.5)
                        )
                        self.thr_var.set(self.threshold)
                        self.thr_value.config(text=f"{self.threshold:.2f}")
                    if self.current is not None:
                        cached = self.score_cache.get(self.current.name)
                        if cached is not None:
                            self.current_score = cached
                        elif self.predictor.available:
                            self.current_score = self.predictor.score(self.current)
                            if self.current_score is not None:
                                self.score_cache[self.current.name] = self.current_score
                        self._update_score_display()

            self.root.after(0, finish)

        threading.Thread(target=worker, daemon=True).start()


def pick_output_device(preferred: str | int | None) -> int | None:
    devices = sd.query_devices()

    def is_output(i):
        return devices[i]["max_output_channels"] > 0

    if isinstance(preferred, int):
        if 0 <= preferred < len(devices) and is_output(preferred):
            return preferred
        print(f"[device] index {preferred} is not a valid output device", file=sys.stderr)
        return None

    if isinstance(preferred, str):
        needle = preferred.lower()
        for i, d in enumerate(devices):
            if is_output(i) and needle in d["name"].lower():
                return i
        print(f"[device] no output device matches '{preferred}'", file=sys.stderr)
        return None

    for i, d in enumerate(devices):
        if is_output(i) and "g433" in d["name"].lower():
            return i
    return None


def main():
    ap = argparse.ArgumentParser(description="Label bark recordings YES/NO.")
    ap.add_argument("--device", help="Output device name substring or index (default: auto-pick G433 headset)")
    ap.add_argument("--list-devices", action="store_true", help="List output devices and exit")
    args = ap.parse_args()

    if args.list_devices:
        for i, d in enumerate(sd.query_devices()):
            if d["max_output_channels"] > 0:
                print(f"[{i}] {d['name']}  out_ch={d['max_output_channels']}  sr={d['default_samplerate']}")
        return

    preferred: str | int | None = args.device
    if isinstance(preferred, str) and preferred.isdigit():
        preferred = int(preferred)
    dev = pick_output_device(preferred)
    if dev is not None:
        sd.default.device = (None, dev)
        print(f"[device] using output [{dev}] {sd.query_devices(dev)['name']}")
    else:
        print(f"[device] using system default ({sd.default.device})")

    if not REC_DIR.is_dir():
        print(f"Recordings folder not found: {REC_DIR}", file=sys.stderr)
        sys.exit(1)
    all_files = sorted(REC_DIR.glob("*.wav"))
    files = [p for p in all_files if bark_id_from_name(p.name).split("_")[0] not in EXCLUDE_DATES]
    excluded = len(all_files) - len(files)
    if excluded:
        print(f"[filter] excluded {excluded} recordings from dates {sorted(EXCLUDE_DATES)}")
    if not files:
        print("No .wav files found.", file=sys.stderr)
        sys.exit(1)

    db_map = parse_log_dbfs()
    wb = LabelerWorkbook(XLSX)
    predictor = BarkPredictor()
    if predictor.available:
        print(f"[model] loaded bark_cnn.pt  (OOF acc={predictor.meta.get('oof_accuracy'):.3f})")
    else:
        print(f"[model] unavailable: {predictor.error}")

    root = tk.Tk()
    App(root, files, db_map, wb, predictor)
    root.mainloop()


if __name__ == "__main__":
    main()
